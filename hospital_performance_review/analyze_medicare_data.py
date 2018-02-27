#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jul  7 18:23:30 2017

@author: zan
"""

"""
    Healthcare Analytics – 
    Recommender system for hospitals based on Medicare ratings and patient surveys
"""

# Download the Medicare Hospital Compare data from the website
import requests
import os
import zipfile
import openpyxl
import pandas as pd
import glob
import sqlite3


url = 'https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip'
r = requests.get(url)

# Create the staging directory
staging_dir = "staging"
os.mkdir(staging_dir)

# Confirm the staging directory path
os.path.isdir(staging_dir)

# Machine independent path to create files
zip_file = os.path.join(staging_dir, "Hospital_Revised_Flatfiles.zip")

# Write the files to the computer
zf = open(zip_file,"wb")
zf.write(r.content)
zf.close()

# Program to unzip the files
z = zipfile.ZipFile(zip_file,"r")
z.extractall(staging_dir)
z.close()

# Delete the corrupt csv
os.listdir(os.getcwd())

os.remove("staging/FY2015_Percent_Change_in_Medicare_Payments.csv")

# Download the proprietary in-house information
k_url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"
r = requests.get(k_url)

# Open the excel files with the proprietary in-house information
xf = open("hospital_ranking_focus_states.xlsx","wb")
xf.write(r.content)
xf.close()

# Read the proprietary excel files
wb = openpyxl.load_workbook("hospital_ranking_focus_states.xlsx")

# Write the rows of the hospital_ranking sheet
hospital_national_ranking = wb.get_sheet_by_name("Hospital National Ranking")

i = 1
while hospital_national_ranking.cell(row=i, column = 1).value != None:
    print(hospital_national_ranking.cell(row=i, column = 1).value, "|", hospital_national_ranking.cell(row = i, column = 2).value)
    i += 1

# Write the rows of the focus_states sheet
focus_states = wb.get_sheet_by_name("Focus States")

i = 1
while focus_states.cell(row=i, column=1).value != None:
    print(focus_states.cell(row=i, column=1).value, "|", focus_states.cell(row = i, column = 2).value)
    i += 1

# Convert the proprietary information to Pandas
hospital_national_ranking = pd.DataFrame(hospital_national_ranking.values, dtype = str)
hospital_national_ranking.columns = ['provider_id','ranking']

focus_states = pd.DataFrame(focus_states.values, dtype = str)
focus_states.columns = ['state_name','state_abbreviation']

# Convert all letters to lower case
hospital_national_ranking.columns = hospital_national_ranking.columns.str.lower()
focus_states.columns = focus_states.columns.str.lower()

# Drop the first row containing the old column headers
hospital_national_ranking.drop(hospital_national_ranking.head(1).index, inplace=True)
focus_states.drop(focus_states.head(1).index, inplace = True)

# Write the excel files to the staging directory
hospital_national_ranking.to_csv(os.path.join(staging_dir,r'hospital_national_ranking.csv'), encoding = 'utf-8')
focus_states.to_csv(os.path.join(staging_dir,r'focus_states.csv'), encoding = 'utf-8')

# Format the file names
"""
Convert all letters to lower case DONE
Replace each blank “ “ with an underscore “_” DONE
Replace each dash or hyphen “-“ with an underscore “_” DONE
Replace each percent sign “%” with the string “pct” DONE
Replace each forward slash “/” with an underscore “_” DONE
If a table name starts with anything other than a letter “a” through “z” then prepend “t_” to the front of the table name
If a column name starts with anything other than a letter “a” through “z” then prepend “c_” to the front of the column name
"""
os.listdir("staging")

# Seperate filename from extension
sep = os.sep

# Change the casing
for n in os.listdir("staging"):
    print(n)
    if os.path.isfile("staging" + sep + n):
        filename_one, extension = os.path.splitext(n)
        os.rename("staging" + sep + n, "staging" + sep + filename_one.lower() + extension)

# Remove the blanks, -, %, and /
for n in os.listdir("staging"):
    print (n)
    if os.path.isfile("staging" + sep + n):
        filename_zero, extension = os.path.splitext(n)
        os.rename("staging" + sep + n , "staging" + sep + filename_zero.replace(' ','_').replace('-','_').replace('%','pct').replace('/','_') + extension)

# Show the new file names
print ('\n--------------------------------\n')
for n in os.listdir("staging"):
    print (n)

"""
In order to fix all of the column headers and to solve the encoding issues and remove nulls, 
first read in all of the CSV's to python as dataframes, then make changes and rewrite the old files
"""
files = glob.glob(os.path.join("staging" + "/*.csv"))

print(files)

# Create an empty dictionary to hold the dataframes from csvs
dict_ = {}

# Write the files into the dictionary
for file in files:
    fname = os.path.basename(file)
    fname = fname.replace('.csv', '')
    dict_[fname] = pd.read_csv(file, header = 0, dtype = str, encoding = 'cp1252').fillna('')

"""
Convert all letters to lower case 
Replace each blank “ “ with an underscore “_” 
Replace each dash or hyphen “-“ with an underscore “_” 
Replace each percent sign “%” with the string “pct” 
Replace each forward slash “/” with an underscore “_” 
If a table name starts with anything other than a letter “a” through “z” then prepend “t_” to the front of the table name
If a column name starts with anything other than a letter “a” through “z” then prepend “c_” to the front of the column name
"""
# Convert all letters to lower case
for file in files:
    fname = os.path.basename(file)
    fname = fname.replace('.csv', '')
    dict_[fname].columns = dict_[fname].columns.str.lower()

# Fix the column headers 
for file in files:
    fname = os.path.basename(file)
    fname = fname.replace('.csv', '')
    dict_[fname].rename(columns = lambda x: x.replace(' ', '_').replace('-', '_').replace('%', 'pct').replace('/', '_'), inplace = True)
    # Add the prefix to columns beginning with non-alpha characters
    mask = dict_[fname].columns.str[0].str.isalpha()
    dict_[fname].columns = dict_[fname].columns.where(mask, 'c_' + dict_[fname].columns) 

"""
Rewrite the dataframes to csv in utf8, overwriting old values 

for file in dict_:
    dict_[file].to_csv(file, encoding = 'utf-8')
"""

# Create the SQL Lite database
conn = sqlite3.connect("medicare_hospital_compare.db")

# Convert the dict_[file]'s to SQL tables
for key, df in dict_.items(): 
    df.to_sql(key, conn, flavor = None, schema = None, if_exists = 'replace', 
              index = False, index_label = None, chunksize = None, dtype = None)

# Queries

cur = conn.cursor()

# Top 100 hospitals nationwide
national_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
group by hospital_national_ranking.ranking
having cast(hospital_national_ranking.ranking as decimal) < 101
order by cast(hospital_national_ranking.ranking as decimal) asc;""", conn)

# Top 100 hospitals California
ca_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%CA%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Florida
fl_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%FL%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Georgia
ga_results = pd.read_sql_query("""select 
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%GA%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Illinois
il_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%IL%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Kansas
ks_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%KS%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Michigan
mi_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%MI%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals New York
ny_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%NY%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Ohio
oh_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%OH%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Pennsylvania
pa_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%PA%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# # Top 100 hospitals Texas
tx_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%TX%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

ranking_results_list = [national_results, ca_results, fl_results, ga_results, il_results, ks_results, mi_results, ny_results, oh_results, pa_results, tx_results]

# Nationwide measure statistics
nationwide_measures = pd.read_sql_query("""select state,
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital;""", conn)

# CA measure statistics
ca_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%CA%";""", conn)

# FL measure statistics
fl_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%FL%";""", conn)

# GA measure statistics
ga_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%GA%";""", conn)

# IL measure statistics
il_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%IL%";""", conn)

# KS measure statistics
ks_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%KS%";""", conn)

# MI measure statistics
mi_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%MI%";""", conn)

# NY measure statistics
ny_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%NY%";""", conn)

# OH measure statistics
oh_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%OH%";""", conn)

# PA measure statistics
pa_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%PA%";""", conn)

# TX measure statistics
tx_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%TX%";""", conn)

measure_results_list = [nationwide_measures, ca_stat, fl_stat, ga_stat, il_stat, ks_stat, mi_stat, ny_stat, oh_stat, pa_stat, tx_stat]

def fix_result(lst):
    fixed_list = []
    measures_list = []
    for result in lst:
        # Remove the non-numeric string values from 'score'
        result1 = result[result['score'].astype(str).str.isdigit()]
        # Change score to numeric
        result1['score'] = pd.to_numeric(result1['score'])
        fixed_list.append(result1)
    for lst in fixed_list:
        # Calculate the measure statistics
        lst_measure_results = (lst.groupby(['measure_id','measure_name'])['score'].agg(['min','max','mean','std'])
         .rename(columns={'measure_id':'Measure ID','measure_name':'Measure Name','min':'Minimum','max':'Maximum','mean':'Average','std':'Standard Deviation'})
         .rename_axis(['Measure ID','Measure Name'])
         .reset_index())
        measures_list.append(lst_measure_results)
    return measures_list

measure_results_list1 = fix_result(measure_results_list)   

measure_statistics = pd.ExcelWriter("measure_statistics.xlsx")

list_sheet_name = ['Nationwide','California', 'Florida', 'Georgia', 'Illinois', 'Kansas', 'Michigan', 'New York', 'Ohio', 'Pennsylvania', 'Texas']
for df, sheetname in zip(measure_results_list1,list_sheet_name):
    df.to_excel(measure_statistics, sheet_name=sheetname)

measure_statistics.save()

hospital_ranking = pd.ExcelWriter("hospital_ranking.xlsx")
for df, sheetname in zip(ranking_results_list,list_sheet_name):
    df.to_excel(hospital_ranking, sheet_name=sheetname)

hospital_ranking.save()

