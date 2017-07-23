#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jul  7 18:23:30 2017

@author: zan
"""

"""
    Healthcare Analytics – 
    Recommender system for hospitals based on Medicare ratings and patient surveys
"""

# Download the Medicare Hospital Compare data from the website
import requests
import os

url = 'https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip'
r = requests.get(url)

# Create the staging directory
staging_dir = "staging"
os.mkdir(staging_dir)

# Confirm the staging directory path
os.path.isdir(staging_dir)

# Machine independent path to create files
zip_file = os.path.join(staging_dir, "Hospital_Revised_Flatfiles.zip")

# Write the files to the computer
zf = open(zip_file,"wb")
zf.write(r.content)
zf.close()

# Program to unzip the files
import zipfile

z = zipfile.ZipFile(zip_file,"r")
z.extractall(staging_dir)
z.close()

# Delete the corrupt csv
os.listdir(os.getcwd())

os.remove("staging/FY2015_Percent_Change_in_Medicare_Payments.csv")

# Download the proprietary in-house information
k_url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"
r = requests.get(k_url)

# Open the excel files with the proprietary in-house information
xf = open("hospital_ranking_focus_states.xlsx","wb")
xf.write(r.content)
xf.close()

# Read the proprietary excel files
import openpyxl

wb = openpyxl.load_workbook("hospital_ranking_focus_states.xlsx")

# Write the rows of the hospital_ranking sheet
hospital_national_ranking = wb.get_sheet_by_name("Hospital National Ranking")

i = 1
while hospital_national_ranking.cell(row=i, column = 1).value != None:
    print(hospital_national_ranking.cell(row=i, column = 1).value, "|", hospital_national_ranking.cell(row = i, column = 2).value)
    i += 1

# Write the rows of the focus_states sheet
focus_states = wb.get_sheet_by_name("Focus States")

i = 1
while focus_states.cell(row=i, column=1).value != None:
    print(focus_states.cell(row=i, column=1).value, "|", focus_states.cell(row = i, column = 2).value)
    i += 1

# Convert the proprietary information to Pandas
import pandas as pd

hospital_national_ranking = pd.DataFrame(hospital_national_ranking.values, dtype = str)
hospital_national_ranking.columns = ['provider_id','ranking']

focus_states = pd.DataFrame(focus_states.values, dtype = str)
focus_states.columns = ['state_name','state_abbreviation']

# Convert all letters to lower case
hospital_national_ranking.columns = hospital_national_ranking.columns.str.lower()
focus_states.columns = focus_states.columns.str.lower()

# Drop the first row containing the old column headers
hospital_national_ranking.drop(hospital_national_ranking.head(1).index, inplace=True)
focus_states.drop(focus_states.head(1).index, inplace = True)

# Write the excel files to the staging directory
hospital_national_ranking.to_csv(os.path.join(staging_dir,r'hospital_national_ranking.csv'), encoding = 'utf-8')
focus_states.to_csv(os.path.join(staging_dir,r'focus_states.csv'), encoding = 'utf-8')

# Format the file names
"""
Convert all letters to lower case DONE
Replace each blank “ “ with an underscore “_” DONE
Replace each dash or hyphen “-“ with an underscore “_” DONE
Replace each percent sign “%” with the string “pct” DONE
Replace each forward slash “/” with an underscore “_” DONE
If a table name starts with anything other than a letter “a” through “z” then prepend “t_” to the front of the table name
If a column name starts with anything other than a letter “a” through “z” then prepend “c_” to the front of the column name
"""
os.listdir("staging")

# Seperate filename from extension
sep = os.sep

# Change the casing
for n in os.listdir("staging"):
    print(n)
    if os.path.isfile("staging" + sep + n):
        filename_one, extension = os.path.splitext(n)
        os.rename("staging" + sep + n, "staging" + sep + filename_one.lower() + extension)
        
# Show the new file names
print ('\n--------------------------------\n')
for n in os.listdir("staging"):
    print (n)

# Remove the blanks, -, %, and /
for n in os.listdir("staging"):
    print (n)
    if os.path.isfile("staging" + sep + n):
        filename_zero, extension = os.path.splitext(n)
        os.rename("staging" + sep + n , "staging" + sep + filename_zero.replace(' ','_').replace('-','_').replace('%','pct').replace('/','_') + extension)

# Show the new file names
print ('\n--------------------------------\n')
for n in os.listdir("staging"):
    print (n)

"""
In order to fix all of the column headers and to solve the encoding issues and remove nulls, 
first read in all of the CSV's to python as dataframes, then make changes and rewrite the old files
"""
import os
import glob

files = glob.glob(os.path.join("staging" + "/*.csv"))

print(files)

# Create an empty dictionary to hold the dataframes from csvs
dict_ = {}

# Write the files into the dictionary
for file in files:
    fname = os.path.basename(file)
    fname = fname.replace('.csv', '')
    dict_[fname] = pd.read_csv(file, header = 0, dtype = str, encoding = 'cp1252').fillna('')

"""
Convert all letters to lower case 
Replace each blank “ “ with an underscore “_” 
Replace each dash or hyphen “-“ with an underscore “_” 
Replace each percent sign “%” with the string “pct” 
Replace each forward slash “/” with an underscore “_” 
If a table name starts with anything other than a letter “a” through “z” then prepend “t_” to the front of the table name
If a column name starts with anything other than a letter “a” through “z” then prepend “c_” to the front of the column name
"""
# Convert all letters to lower case
for file in files:
    fname = os.path.basename(file)
    fname = fname.replace('.csv', '')
    dict_[fname].columns = dict_[fname].columns.str.lower()

# Fix the column headers 
for file in files:
    fname = os.path.basename(file)
    fname = fname.replace('.csv', '')
    dict_[fname].rename(columns = lambda x: x.replace(' ', '_'), inplace = True)

for file in files:
    fname = os.path.basename(file)
    fname = fname.replace('.csv', '')
    dict_[fname].rename(columns = lambda x: x.replace('-', '_'), inplace = True)

for file in files:
    fname = os.path.basename(file)
    fname = fname.replace('.csv', '')
    dict_[fname].rename(columns = lambda x: x.replace('%', 'pct'), inplace = True)

for file in files:
    fname = os.path.basename(file)
    fname = fname.replace('.csv', '')
    dict_[fname].rename(columns = lambda x: x.replace('/', '_'), inplace = True)

# Add the prefix to columns beginning with non-alpha characters
for file in files:
    fname = os.path.basename(file)
    fname = fname.replace('.csv', '')
    mask = dict_[fname].columns.str[0].str.isalpha()
    dict_[fname].columns = dict_[fname].columns.where(mask, 'c_' + dict_[fname].columns) 

"""
Rewrite the dataframes to csv in utf8, overwriting old values 

for file in dict_:
    dict_[file].to_csv(file, encoding = 'utf-8')
"""

# Create the SQL Lite database
import sqlite3

conn = sqlite3.connect("medicare_hospital_compare.db")

# Convert the dict_[file]'s to SQL tables
for key, df in dict_.items(): 
    df.to_sql(key, conn, flavor = None, schema = None, if_exists = 'replace', 
              index = False, index_label = None, chunksize = None, dtype = None)

# Queries

cur = conn.cursor()

# Top 100 hospitals nationwide
national_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
group by hospital_national_ranking.ranking
having cast(hospital_national_ranking.ranking as decimal) < 101
order by cast(hospital_national_ranking.ranking as decimal) asc;""", conn)

# Top 100 hospitals California
ca_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%CA%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Florida
fl_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%FL%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Georgia
ga_results = pd.read_sql_query("""select 
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%GA%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Illinois
il_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%IL%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Kansas
ks_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%KS%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Michigan
mi_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%MI%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals New York
ny_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%NY%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Ohio
oh_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%OH%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Top 100 hospitals Pennsylvania
pa_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%PA%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# # Top 100 hospitals Texas
tx_results = pd.read_sql_query("""select
		  hospital_general_information.provider_id as 'Provider ID',
		  hospital_name as 'Hospital Name',
          city as 'City',
          state as 'State',
          county_name as 'County'
from hospital_general_information left join hospital_national_ranking
on hospital_general_information.provider_id = hospital_national_ranking.provider_id
where state like '%TX%'
group by hospital_national_ranking.ranking
order by cast(hospital_national_ranking.ranking as decimal) asc
limit 100;""", conn)

# Nationwide measure statistics
nationwide_measures = pd.read_sql_query("""select state,
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital;""", conn)

# Remove the non-numeric string values from 'score'
nationwide_measures1 = nationwide_measures[nationwide_measures['score'].astype(str).str.isdigit()]

# Change score to numeric
nationwide_measures1['score'] = pd.to_numeric(nationwide_measures1['score'])

# Calculate the measure statistics
nationwide_measure_results = (nationwide_measures1.groupby(['measure_id','measure_name'])['score'].agg(['min','max','mean','std'])
         .rename(columns={'measure_id':'Measure ID','measure_name':'Measure Name','min':'Minimum','max':'Maximum','mean':'Average','std':'Standard Deviation'})
         .rename_axis(['Measure ID','Measure Name'])
         .reset_index())

# CA measure statistics
ca_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%CA%";""", conn)

# Remove the non-numeric string values from 'score'
ca_stat1 = ca_stat[ca_stat['score'].astype(str).str.isdigit()]

# Change score to numeric
ca_stat1['score'] = pd.to_numeric(ca_stat1['score'])

# Calculate the measure statistics
ca_measure_results = (ca_stat1.groupby(['measure_id','measure_name'])['score'].agg(['min','max','mean','std'])
         .rename(columns={'measure_id':'Measure ID','measure_name':'Measure Name','min':'Minimum','max':'Maximum','mean':'Average','std':'Standard Deviation'})
         .rename_axis(['Measure ID','Measure Name'])
         .reset_index())

# FL measure statistics
fl_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%FL%";""", conn)

# Remove the non-numeric string values from 'score'
fl_stat1 = fl_stat[fl_stat['score'].astype(str).str.isdigit()]

# Change score to numeric
fl_stat1['score'] = pd.to_numeric(fl_stat1['score'])

# Calculate the measure statistics
fl_measure_results = (fl_stat1.groupby(['measure_id','measure_name'])['score'].agg(['min','max','mean','std'])
         .rename(columns={'measure_id':'Measure ID','measure_name':'Measure Name','min':'Minimum','max':'Maximum','mean':'Average','std':'Standard Deviation'})
         .rename_axis(['Measure ID','Measure Name'])
         .reset_index())

# GA measure statistics
ga_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%GA%";""", conn)

# Remove the non-numeric string values from 'score'
ga_stat1 = ga_stat[ga_stat['score'].astype(str).str.isdigit()]

# Change score to numeric
ga_stat1['score'] = pd.to_numeric(ga_stat1['score'])

# Calculate the measure statistics
ga_measure_results = (ga_stat1.groupby(['measure_id','measure_name'])['score'].agg(['min','max','mean','std'])
         .rename(columns={'measure_id':'Measure ID','measure_name':'Measure Name','min':'Minimum','max':'Maximum','mean':'Average','std':'Standard Deviation'})
         .rename_axis(['Measure ID','Measure Name'])
         .reset_index())

# IL measure statistics
il_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%IL%";""", conn)

# Remove the non-numeric string values from 'score'
il_stat1 = il_stat[il_stat['score'].astype(str).str.isdigit()]

# Change score to numeric
il_stat1['score'] = pd.to_numeric(il_stat1['score'])

# Calculate the measure statistics
il_measure_results = (il_stat1.groupby(['measure_id','measure_name'])['score'].agg(['min','max','mean','std'])
         .rename(columns={'measure_id':'Measure ID','measure_name':'Measure Name','min':'Minimum','max':'Maximum','mean':'Average','std':'Standard Deviation'})
         .rename_axis(['Measure ID','Measure Name'])
         .reset_index())

# KS measure statistics
ks_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%KS%";""", conn)

# Remove the non-numeric string values from 'score'
ks_stat1 = ks_stat[ks_stat['score'].astype(str).str.isdigit()]

# Change score to numeric
ks_stat1['score'] = pd.to_numeric(ks_stat1['score'])

# Calculate the measure statistics
ks_measure_results = (ks_stat1.groupby(['measure_id','measure_name'])['score'].agg(['min','max','mean','std'])
         .rename(columns={'measure_id':'Measure ID','measure_name':'Measure Name','min':'Minimum','max':'Maximum','mean':'Average','std':'Standard Deviation'})
         .rename_axis(['Measure ID','Measure Name'])
         .reset_index())

# MI measure statistics
mi_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%MI%";""", conn)

# Remove the non-numeric string values from 'score'
mi_stat1 = mi_stat[mi_stat['score'].astype(str).str.isdigit()]

# Change score to numeric
mi_stat1['score'] = pd.to_numeric(mi_stat1['score'])

# Calculate the measure statistics
mi_measure_results = (mi_stat1.groupby(['measure_id','measure_name'])['score'].agg(['min','max','mean','std'])
         .rename(columns={'measure_id':'Measure ID','measure_name':'Measure Name','min':'Minimum','max':'Maximum','mean':'Average','std':'Standard Deviation'})
         .rename_axis(['Measure ID','Measure Name'])
         .reset_index())

# NY measure statistics
ny_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%NY%";""", conn)

# Remove the non-numeric string values from 'score'
ny_stat1 = ny_stat[ny_stat['score'].astype(str).str.isdigit()]

# Change score to numeric
ny_stat1['score'] = pd.to_numeric(ny_stat1['score'])

# Calculate the measure statistics
ny_measure_results = (ny_stat1.groupby(['measure_id','measure_name'])['score'].agg(['min','max','mean','std'])
         .rename(columns={'measure_id':'Measure ID','measure_name':'Measure Name','min':'Minimum','max':'Maximum','mean':'Average','std':'Standard Deviation'})
         .rename_axis(['Measure ID','Measure Name'])
         .reset_index())

# OH measure statistics
oh_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%OH%";""", conn)

# Remove the non-numeric string values from 'score'
oh_stat1 = oh_stat[oh_stat['score'].astype(str).str.isdigit()]

# Change score to numeric
oh_stat1['score'] = pd.to_numeric(oh_stat1['score'])

# Calculate the measure statistics
oh_measure_results = (oh_stat1.groupby(['measure_id','measure_name'])['score'].agg(['min','max','mean','std'])
         .rename(columns={'measure_id':'Measure ID','measure_name':'Measure Name','min':'Minimum','max':'Maximum','mean':'Average','std':'Standard Deviation'})
         .rename_axis(['Measure ID','Measure Name'])
         .reset_index())

# PA measure statistics
pa_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%PA%";""", conn)

# Remove the non-numeric string values from 'score'
pa_stat1 = pa_stat[pa_stat['score'].astype(str).str.isdigit()]

# Change score to numeric
pa_stat1['score'] = pd.to_numeric(pa_stat1['score'])

# Calculate the measure statistics
pa_measure_results = (pa_stat1.groupby(['measure_id','measure_name'])['score'].agg(['min','max','mean','std'])
         .rename(columns={'measure_id':'Measure ID','measure_name':'Measure Name','min':'Minimum','max':'Maximum','mean':'Average','std':'Standard Deviation'})
         .rename_axis(['Measure ID','Measure Name'])
         .reset_index())

# TX measure statistics
tx_stat = pd.read_sql_query("""select 
		  measure_id,
		  measure_name,
		  score
from timely_and_effective_care___hospital
where state like "%TX%";""", conn)

# Remove the non-numeric string values from 'score'
tx_stat1 = tx_stat[tx_stat['score'].astype(str).str.isdigit()]

# Change score to numeric
tx_stat1['score'] = pd.to_numeric(tx_stat1['score'])

# Calculate the measure statistics
tx_measure_results = (tx_stat1.groupby(['measure_id','measure_name'])['score'].agg(['min','max','mean','std'])
         .rename(columns={'measure_id':'Measure ID','measure_name':'Measure Name','min':'Minimum','max':'Maximum','mean':'Average','std':'Standard Deviation'})
         .rename_axis(['Measure ID','Measure Name'])
         .reset_index())
"""
“hospital_ranking.xlsx” workbook in the local directory without using any path names.
It should have a first sheet named “Nationwide”. 
It should have the following column headers “Provider ID”, “Hospital Name”, “City”, “State”, and “County”. 
Follow this header row with the top 100 hospitals as ranked by the in house proprietary system, 
ordered by rank. For the state column, the data should use the 2 letter state abbreviation.
"""

"""
pd.ExcelWriter

hospital_ranking = pd.ExcelWriter("hospital_ranking.xlsx")

list_dfs = [national_results,ca_results,fl_results,ga_results,il_results,ks_results,mi_results,ny_results,oh_results,pa_results,tx_results]
list_sheet_name = ['Nationwide','California', 'Florida', 'Georgia', 'Illinois', 'Kansas', 'Michigan', 'New York', 'Ohio', 'Pennsylvania', 'Texas']
for df, sheetname in zip(list_dfs,list_sheet_name):
    df.to_excel(hospital_ranking, sheet_name=sheetname)

hospital_ranking.save()
"""

from openpyxl import Workbook

# Create the hospital_ranking workbook
hospital_ranking = Workbook()
dest_filename1 = "hospital_ranking.xlsx"

ws1 = hospital_ranking.active
ws1.title = "Nationwide"

from openpyxl.utils.dataframe import dataframe_to_rows

# Write the nationwide query to ws1
for r in dataframe_to_rows(national_results, index = False, header = True):
    ws1.append(r)

for cell in ws1['A'] + ws1[1]:
    cell.style = 'Pandas'

# Create the worksheet for each focus state

# CA
ws2 = hospital_ranking.create_sheet(title = 'California')
ws2 = hospital_ranking.get_sheet_by_name('California')

# Write the CA query to ws2
for r in dataframe_to_rows(ca_results, index = False, header = True):
    ws2.append(r)

for cell in ws2['A'] + ws2[1]:
    cell.style = 'Pandas'

# FL
ws3 = hospital_ranking.create_sheet(title = 'Florida')
ws3 = hospital_ranking.get_sheet_by_name('Florida')


# Write the FL query to ws3
for r in dataframe_to_rows(fl_results, index = False, header = True):
    ws3.append(r)

for cell in ws3['A'] + ws3[1]:
    cell.style = 'Pandas'

# GA
ws4 = hospital_ranking.create_sheet(title = 'Georgia')
ws4 = hospital_ranking.get_sheet_by_name('Georgia')


# Write the GA query to ws4
for r in dataframe_to_rows(ga_results, index = False, header = True):
    ws4.append(r)

for cell in ws4['A'] + ws4[1]:
    cell.style = 'Pandas'

# IL
ws5 = hospital_ranking.create_sheet(title = 'Illinois')
ws5 = hospital_ranking.get_sheet_by_name('Illinois')


# Write the IL query to ws5
for r in dataframe_to_rows(il_results, index = False, header = True):
    ws5.append(r)

for cell in ws5['A'] + ws5[1]:
    cell.style = 'Pandas'

# KS
ws6 = hospital_ranking.create_sheet(title = 'Kansas')
ws6 = hospital_ranking.get_sheet_by_name('Kansas')


# Write the KS query to ws6
for r in dataframe_to_rows(ks_results, index = False, header = True):
    ws6.append(r)

for cell in ws6['A'] + ws6[1]:
    cell.style = 'Pandas'

# MI
ws7 = hospital_ranking.create_sheet(title = 'Michigan')
ws7 = hospital_ranking.get_sheet_by_name('Michigan')


# Write the MI query to ws7
for r in dataframe_to_rows(mi_results, index = False, header = True):
    ws7.append(r)

for cell in ws7['A'] + ws7[1]:
    cell.style = 'Pandas'

# NY
ws8 = hospital_ranking.create_sheet(title = 'New York')
ws8 = hospital_ranking.get_sheet_by_name('New York')


# Write the NY query to ws8
for r in dataframe_to_rows(ny_results, index = False, header = True):
    ws8.append(r)

for cell in ws8['A'] + ws8[1]:
    cell.style = 'Pandas'

# OH
ws9 = hospital_ranking.create_sheet(title = 'Ohio')
ws9 = hospital_ranking.get_sheet_by_name('Ohio')


# Write the OH query to ws9
for r in dataframe_to_rows(oh_results, index = False, header = True):
    ws9.append(r)

for cell in ws9['A'] + ws9[1]:
    cell.style = 'Pandas'

# PA
ws10 = hospital_ranking.create_sheet(title = 'Pennsylvania')
ws10 = hospital_ranking.get_sheet_by_name('Pennsylvania')


# Write the PA query to ws10
for r in dataframe_to_rows(pa_results, index = False, header = True):
    ws10.append(r)

for cell in ws10['A'] + ws10[1]:
    cell.style = 'Pandas'

# TX
ws11 = hospital_ranking.create_sheet(title = 'Texas')
ws11 = hospital_ranking.get_sheet_by_name('Texas')


# Write the TX query to ws11
for r in dataframe_to_rows(tx_results, index = False, header = True):
    ws11.append(r)

for cell in ws11['A'] + ws11[1]:
    cell.style = 'Pandas'

hospital_ranking.save(filename = dest_filename1)


"""
create a hospital ranking MS Excel Workbook named “measures_statistics.xlsx” in the local directory without using any path names.
From the table timely_and_effective___hospital query out the state, measure_id, measure_name, and score. 
It should have a first sheet named “Nationwide”. It should have the following column headers “Measure ID”, “Measure Name”, “Minimum”, “Maximum”, “Average”, and “Standard Deviation”. 
Sort by measure_id. Calculate the minimum, maximum, average, and standard deviation for that measure for all hospitals nationwide.
For each of the states in the focus list, it should have a separate sheet for each state.
"""

# Create the measure_statistics workbook
measures_statistics = Workbook()
dest_filename2 = "measures_statistics.xlsx"

ws12 = measures_statistics.active
ws12.title = "Nationwide"

for r in dataframe_to_rows(nationwide_measure_results, index = False, header = True):
    ws12.append(r)

for cell in ws12['A'] + ws12[1]:
    cell.style = 'Pandas'

# Create the worksheet for each focus state
# CA
ws13 = measures_statistics.create_sheet(title = 'California')
ws13 = measures_statistics.get_sheet_by_name('California')

# Write the measure results 
for r in dataframe_to_rows(ca_measure_results, index = False, header = True):
    ws13.append(r)

for cell in ws13['A'] + ws13[1]:
    cell.style = 'Pandas'

# FL
ws14 = measures_statistics.create_sheet(title = 'Florida')
ws14 = measures_statistics.get_sheet_by_name('Florida')

# Write the measure results 
for r in dataframe_to_rows(fl_measure_results, index = False, header = True):
    ws14.append(r)

for cell in ws14['A'] + ws14[1]:
    cell.style = 'Pandas'

# GA
ws15 = measures_statistics.create_sheet(title = 'Georgia')
ws15 = measures_statistics.get_sheet_by_name('Georgia')

# Write the measure results 
for r in dataframe_to_rows(ga_measure_results, index = False, header = True):
    ws15.append(r)

for cell in ws15['A'] + ws15[1]:
    cell.style = 'Pandas'

# IL
ws16 = measures_statistics.create_sheet(title = 'Illinois')
ws16 = measures_statistics.get_sheet_by_name('Illinois')

# Write the measure results 
for r in dataframe_to_rows(il_measure_results, index = False, header = True):
    ws16.append(r)

for cell in ws16['A'] + ws16[1]:
    cell.style = 'Pandas'

#KS
ws17 = measures_statistics.create_sheet(title = 'Kansas')
ws17 = measures_statistics.get_sheet_by_name('Kansas')

# Write the measure results 
for r in dataframe_to_rows(ks_measure_results, index = False, header = True):
    ws17.append(r)

for cell in ws17['A'] + ws17[1]:
    cell.style = 'Pandas'

# MI
ws18 = measures_statistics.create_sheet(title = "Michigan")
ws18 = measures_statistics.get_sheet_by_name('Michigan')

# Write the measure results 
for r in dataframe_to_rows(mi_measure_results, index = False, header = True):
    ws18.append(r)

for cell in ws18['A'] + ws18[1]:
    cell.style = 'Pandas'

# NY
ws19 = measures_statistics.create_sheet(title = 'New York')
ws19 = measures_statistics.get_sheet_by_name('New York')

# Write the measure results 
for r in dataframe_to_rows(ny_measure_results, index = False, header = True):
    ws19.append(r)

for cell in ws19['A'] + ws19[1]:
    cell.style = 'Pandas'

# OH
ws20 = measures_statistics.create_sheet(title = 'Ohio')
ws20 = measures_statistics.get_sheet_by_name('Ohio')

# Write the measure results 
for r in dataframe_to_rows(oh_measure_results, index = False, header = True):
    ws20.append(r)

for cell in ws20['A'] + ws20[1]:
    cell.style = 'Pandas'

# PA
ws21 = measures_statistics.create_sheet(title = 'Pennsylvania')
ws21 = measures_statistics.get_sheet_by_name('Pennsylvania')

# Write the measure results 
for r in dataframe_to_rows(pa_measure_results, index = False, header = True):
    ws21.append(r)

for cell in ws21['A'] + ws21[1]:
    cell.style = 'Pandas'

#TX
ws22 = measures_statistics.create_sheet(title = 'Texas')
ws22 = measures_statistics.get_sheet_by_name('Texas')

# Write the measure results 
for r in dataframe_to_rows(tx_measure_results, index = False, header = True):
    ws22.append(r)

for cell in ws22['A'] + ws22[1]:
    cell.style = 'Pandas'

# Save the measure_statistics workbook
measures_statistics.save(filename = dest_filename2)

